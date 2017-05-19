# -*- coding: utf-8 -*-

from openpyxl import load_workbook

def readOne():
    wb = load_workbook(filename="py.xlsx")

    sheets = wb.get_sheet_names()
    sheet0 = sheets[0] # 第一个表格的名称
    ws = wb.get_sheet_by_name('Sheet1')

    rows = ws.rows
    columns = ws.columns

    content = []
    for row in rows:
        line = [col.value for col in row]
        print(line)
        content.append(line)

    print(ws.cell('B12').value)
    print(ws.cell(row=12,column=2).value)

    prev_row = [None for i in range(sheet0.ncols)]
    for row_index in range(sheet0.nrows):
        row= []
        for col_index in range(sheet0.ncols):
            value = sheet0.cell(rowx=row_index,colx=col_index).value
            if len(value) == 0:
                value = prev_row[col_index]
            row.append(value)
        prev_row = row
        content.append(row)
        print('新的写法：' ,row)

        Exception
readOne()